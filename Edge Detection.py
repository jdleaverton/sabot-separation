# -*- coding: utf-8 -*-
"""
Created on Thu Jan 21 11:27:35 2021

@author: JD
"""

from tkinter.filedialog import askdirectory
import matplotlib.pyplot as plt
import matplotlib as mpl
import numpy as np
import glob
import cv2
from math import sqrt
from os.path import basename
from openpyxl import Workbook

# set image directory for sabot stripper pictures
imageFolder = askdirectory(title='Select Folder to import images')

imageLocation = imageFolder + '/*.png'

# create excel sheet
wb = Workbook()
ws = wb.active
def createSheetTitles():
    ws.title = 'Distance'
    ws['A1'] = 'Image Name'
    ws['B1'] = 'Average Min Distance'
    ws['C1'] = 'Average Centroid Distance'
    ws['D1'] = 'Average Max Distance'
    ws['E1'] = 'Min Dist 1'
    ws['F1'] = 'Min Dist 2'
    ws['G1'] = 'Min Dist 3'
    ws['H1'] = 'Min Dist 4'
    ws['I1'] = 'Centroid Dist 1'
    ws['J1'] = 'Centroid Dist 2'
    ws['K1'] = 'Centroid Dist 3'
    ws['L1'] = 'Centroid Dist 4'
    ws['M1'] = 'Max Dist 1'
    ws['N1'] = 'Max Dist 2'
    ws['O1'] = 'Max Dist 3'
    ws['P1'] = 'Max Dist 4'
createSheetTitles()

def average(lst):
    return sum(lst) / len(lst)

def pixels2inches(pixels):
    return pixels / 306

def pixels2mm(pixels):
    return (pixels * 25.4) / 306

iterations = 0
# for loop opening each file and performing operations
for filename in glob.glob(imageLocation):
    img = cv2.imread(filename)
    imageNameList = basename(filename).split('.')
    imageName = imageNameList[0]
    
    
    
    
    # optional resize
    center = 1000
    #dsize = (center, center)
    #img = cv2.resize(img, dsize)
    
    # convert the image to grayscale
    gray_image = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # convert the grayscale image to binary image
    ret,thresh = cv2.threshold(gray_image,127,255,0)

    # find contours in the binary image
    contours, hierarchy = cv2.findContours(thresh, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    
    # initialize arrays for min, max, centroid distances for all petals on one image
    minDistance = []
    maxDistance = []
    centroidDistance = []
    circularity = []
    verticies = []
    # iterate through contours
    for c in contours:
        
        
        # gets rid of the program thinking the edge of the picture is a contour
        if cv2.arcLength(c, True) > 0.75 * img.shape[0]:
            continue

        
        area = cv2.contourArea(c)
        perimeter = cv2.arcLength(c, True)
        
        #circularityV = perimeter**2 / (4 * np.pi * area)
        #circularity.append(circularityV)
        approx = cv2.approxPolyDP(c, 0.005 * perimeter, True)
        verticies.append((len(approx)))
        
        # initialize array that edge points are added to to find min and max of them
        distance = []
        # iterate through edge points of each contour
        for i in range(len(c)):
            
            xCordinate = c[i, 0, 0]
            yCordinate = c[i, 0, 1]
            
            # distance 
            distanceToCenter = sqrt( (xCordinate - center)**2 + (yCordinate - center)**2 )
            distance.append(pixels2mm(distanceToCenter))
        
        # find distance min and max of edge points for each contour
        minDistance.append(min(distance))
        maxDistance.append(max(distance))
        
        
        ''' centroid calculations '''
        
        # calculate moments for each contour
        M = cv2.moments(c)

       # calculate x,y coordinate of center
        cX = int(M["m10"] / M["m00"])
        cY = int(M["m01"] / M["m00"])
        
        centroidDistance.append(pixels2mm(sqrt( (cX - center)**2 + (cY - center)**2)))
        
        # drawing on image
        #cv2.line(img, (center, center), (cX, cY), (255, 0, 0))
        cv2.drawContours(img, c, -1, (0, 255, 0), 10)
        cv2.drawContours(img, [approx], 0, (0), 5)
        cv2.circle(img, (cX, cY), 8, (255, 255, 255), -1)
        #cv2.putText(img, "centroid", (cX - 25, cY - 25),cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 2)
        
        #display the image
        #if iterations <= 3:
            #cv2.imshow("Image", img)
            #cv2.imwrite('Image.png', img)
            #cv2.waitKey(0)
    
    '''    
    print('min')
    print(minDistance)
    print('max')
    print(maxDistance)
    print('centroid')
    '''
    print(imageName)
    print(centroidDistance)
    
    #ws.cell(row = iterations + 2, column =5).value = average(circularity)
    ws.cell(row = iterations + 2, column =1).value = imageName
    ws.cell(row = iterations + 2, column =2).value = average(minDistance)
    ws.cell(row = iterations + 2, column =3).value = average(centroidDistance)
    ws.cell(row = iterations + 2, column =4).value = average(maxDistance)
    #ws.cell(row = iterations + 2, column =8).value = average(circularity)
    ws.cell(row = iterations + 2, column =9).value = average(verticies)
    iterations += 1

wb.save('data1.xlsx')
print('Finished.')
    
    


